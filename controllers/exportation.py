
from flask import Blueprint, session, Response, request
import csv
import io
import os
from datetime import datetime
from controllers.auth import login_required
from sqlalchemy import text
from database import engine

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from services.permission_service import permission_required


exportation_routes = Blueprint('exportation', __name__)



def get_ecritures(date_debut=None, date_fin=None, societe_id=None, compte_id=None):
    sql = """
        WITH lignes AS (
            SELECT
                e.id,
                e.date_ecriture,
                e.journal AS piece,
                e.libelle,
                e.montant_ttc AS debit,
                0::numeric AS credit,
                e.societe_id,
                e.compte_debit AS compte_numero
            FROM ecritures_premium e

            UNION ALL

            SELECT
                e.id,
                e.date_ecriture,
                e.journal AS piece,
                e.libelle,
                0::numeric AS debit,
                e.montant_ttc AS credit,
                e.societe_id,
                e.compte_credit AS compte_numero
            FROM ecritures_premium e
        )
        SELECT
            l.id,
            l.date_ecriture,
            l.piece,
            l.libelle,
            l.debit,
            l.credit,
            s.nom AS societe,
            l.compte_numero,
            COALESCE(p.libelle, '') AS compte_libelle
        FROM lignes l
        LEFT JOIN societes_clientes_premium s ON s.id = l.societe_id
        LEFT JOIN plan_comptable p
            ON p.numero = l.compte_numero
           AND (p.societe_id = l.societe_id OR p.societe_id IS NULL)
        WHERE 1=1
    """

    params = {}

    if date_debut:
        sql += " AND l.date_ecriture >= :date_debut"
        params["date_debut"] = date_debut

    if date_fin:
        sql += " AND l.date_ecriture <= :date_fin"
        params["date_fin"] = date_fin

    if societe_id:
        sql += " AND l.societe_id = :societe_id"
        params["societe_id"] = int(societe_id)

    if compte_id:
        sql += " AND p.id = :compte_id"
        params["compte_id"] = int(compte_id)

    sql += " ORDER BY l.date_ecriture DESC, l.id DESC, l.compte_numero"

    with engine.begin() as conn:
        return conn.execute(text(sql), params).fetchall()


def get_societes():
    with engine.begin() as conn:
        return conn.execute(text("""
            SELECT id, nom
            FROM societes_clientes_premium
            ORDER BY nom
        """)).fetchall()


def get_comptes():
    with engine.begin() as conn:
        return conn.execute(text("""
            SELECT id, numero, libelle
            FROM plan_comptable
            ORDER BY numero
        """)).fetchall()


@exportation_routes.route("/")
@login_required
@permission_required("ACCESS_EXPORT")
def export_home():
    username = session.get("username", "Utilisateur")
    societes = get_societes()
    comptes = get_comptes()

    societes_options = "".join(
        [f'<option value="{s[0]}">{s[1]}</option>' for s in societes]
    )

    comptes_options = "".join(
        [f'<option value="{c[0]}">{c[1]} - {c[2]}</option>' for c in comptes]
    )

    return f"""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <title>Export - ComptaPilot</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="bg-light">
        <div class="container py-5">
            <h1 class="mb-4">Export comptable</h1>
            <p class="lead">Connecté en tant que : <strong>{username}</strong></p>

            <div class="card p-4 mb-4">
                <h4>Filtres d'export</h4>

                <form method="GET">
                    <div class="row mb-3">
                        <div class="col-md-3">
                            <label>Date début</label>
                            <input type="date" name="date_debut" class="form-control">
                        </div>

                        <div class="col-md-3">
                            <label>Date fin</label>
                            <input type="date" name="date_fin" class="form-control">
                        </div>

                        <div class="col-md-3">
                            <label>Société</label>
                            <select name="societe_id" class="form-control">
                                <option value="">Toutes les sociétés</option>
                                {societes_options}
                            </select>
                        </div>

                        <div class="col-md-3">
                            <label>Compte</label>
                            <select name="compte_id" class="form-control">
                                <option value="">Tous les comptes</option>
                                {comptes_options}
                            </select>
                        </div>
                    </div>

                    <button formaction="/exportations/ecritures.csv" class="btn btn-success">
                        Export CSV
                    </button>

                    <button formaction="/exportations/ecritures.xlsx" class="btn btn-primary">
                        Export Excel
                    </button>

                    <button formaction="/exportations/journal.pdf" class="btn btn-danger">
                        Export PDF
                    </button>
                    <button formaction="/exportations/backup-db" class="btn btn-dark">
                        Export complet PostgreSQL
                    </button>
                </form>
            </div>

            <a href="/" class="btn btn-secondary">Retour au tableau de bord</a>
        </div>
    </body>
    </html>
    """


@exportation_routes.route("/ecritures.csv")
@login_required
@permission_required("ACCESS_EXPORT")
def export_ecritures_csv():
    rows = get_ecritures(
        request.args.get("date_debut"),
        request.args.get("date_fin"),
        request.args.get("societe_id"),
        request.args.get("compte_id")
    )

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    writer.writerow([
        "ID", "Date", "Pièce", "Libellé", "Débit", "Crédit",
        "Société", "Compte", "Libellé compte"
    ])

    for row in rows:
        writer.writerow(row)

    csv_data = '\ufeff' + output.getvalue()
    output.close()

    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8-sig",
        headers={
            "Content-Disposition": "attachment; filename=ecritures_comptables.csv"
        }
    )


@exportation_routes.route("/ecritures.xlsx")
@login_required
@permission_required("ACCESS_EXPORT")
def export_ecritures_excel():
    rows = get_ecritures(
        request.args.get("date_debut"),
        request.args.get("date_fin"),
        request.args.get("societe_id"),
        request.args.get("compte_id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ecritures comptables"

    headers = [
        "ID", "Date", "Pièce", "Libellé", "Débit", "Crédit",
        "Société", "Compte", "Libellé compte"
    ]

    ws.append(headers)

    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    for row in ws.iter_rows(min_row=2):
        row[4].number_format = '#,##0.00 €'
        row[5].number_format = '#,##0.00 €'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=ecritures_comptables.xlsx"
        }
    )


@exportation_routes.route("/journal.pdf")
@login_required
@permission_required("ACCESS_EXPORT")
def export_journal_pdf():
    rows = get_ecritures(
        request.args.get("date_debut"),
        request.args.get("date_fin"),
        request.args.get("societe_id"),
        request.args.get("compte_id")
    )

    output = io.BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph("Journal comptable - ComptaPilot", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    data = [[
        "Date", "Pièce", "Libellé", "Débit", "Crédit",
        "Société", "Compte"
    ]]

    total_debit = 0
    total_credit = 0

    for row in rows:
        total_debit += row[4] or 0
        total_credit += row[5] or 0

        data.append([
            row[1] or "",
            row[2] or "",
            row[3] or "",
            f"{row[4] or 0:.2f}",
            f"{row[5] or 0:.2f}",
            row[6] or "",
            row[7] or ""
        ])

    data.append([
        "", "", "TOTAL",
        f"{total_debit:.2f}",
        f"{total_credit:.2f}",
        "", ""
    ])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (4, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))

    elements.append(table)
    doc.build(elements)

    output.seek(0)

    return Response(
        output.read(),
        mimetype="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=journal_comptable.pdf"
        }
    )

@exportation_routes.route("/backup-db")
@login_required
@permission_required("ACCESS_EXPORT")
def backup_db():
    date_backup = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"postgres_export_{date_backup}.csv"

    rows = get_ecritures()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "ID", "Date", "Pièce", "Libellé", "Débit", "Crédit",
        "Société", "Compte", "Libellé compte"
    ])

    for row in rows:
        writer.writerow(row)

    csv_data = "\ufeff" + output.getvalue()
    output.close()

    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8-sig",
        headers={
            "Content-Disposition": f"attachment; filename={backup_filename}"
        }
    )
